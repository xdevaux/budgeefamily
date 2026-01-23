"""
Utilitaires pour générer des exports PDF et Excel
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepTogether
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_LEFT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def add_pdf_header(elements, title, user):
    """Ajoute l'en-tête standard avec logo Budgee Family aux exports PDF"""
    styles = getSampleStyleSheet()

    # Logo/Titre Budgee Family
    logo_style = ParagraphStyle(
        'LogoStyle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.HexColor('#1f77b4'),
        fontName='Helvetica-Bold',
        alignment=TA_LEFT
    )
    logo = Paragraph("BUDGEE FAMILY", logo_style)
    elements.append(logo)
    elements.append(Spacer(1, 5))

    # Ligne de séparation
    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1f77b4'), spaceBefore=3, spaceAfter=10))

    # Titre du document
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=5
    )
    title_text = Paragraph(f"{title} - {user.first_name} {user.last_name or ''}", title_style)
    elements.append(title_text)

    # Date de génération
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#666666'))
    date_text = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", date_style)
    elements.append(date_text)
    elements.append(Spacer(1, 15))


def create_excel_workbook():
    """Crée un nouveau classeur Excel avec des styles de base"""
    wb = Workbook()
    return wb


def style_excel_header(ws, row=1):
    """Applique un style aux en-têtes d'un classeur Excel"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def export_upcoming_renewals_excel(renewals, user):
    """Exporte les prochains renouvellements en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Abonnements"

    # En-tête du document
    ws['A1'] = f"Abonnements : Prochains renouvellements - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Date de renouvellement', 'Abonnement', 'Montant', 'Devise', 'Cycle', 'Jours restants']
    ws.append([])  # Ligne vide
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    now = datetime.now().date()
    for sub in renewals:
        days_until = (sub.next_billing_date - now).days
        ws.append([
            sub.next_billing_date.strftime('%d/%m/%Y'),
            sub.name,
            sub.amount,
            sub.currency,
            sub.billing_cycle,
            days_until
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 20

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_upcoming_renewals_pdf(renewals, user):
    """Exporte les prochains renouvellements en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Abonnements : Prochains renouvellements", user)

    # Table
    data = [['Date', 'Abonnement', 'Montant', 'Cycle', 'Jours restants']]
    now = datetime.now().date()

    for sub in renewals:
        days_until = (sub.next_billing_date - now).days
        data.append([
            sub.next_billing_date.strftime('%d/%m/%Y'),
            sub.name[:30],  # Limiter la longueur
            f"{sub.amount:.2f} {sub.currency}",
            sub.billing_cycle,
            str(days_until)
        ])

    table = Table(data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_category_distribution_excel(category_data, user):
    """Exporte la répartition par catégorie en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Répartition par catégorie"

    # En-tête du document
    ws['A1'] = f"Répartition par catégorie - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Catégorie', 'Nombre d\'abonnements', 'Montant mensuel', 'Pourcentage']
    ws.append([])
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    total = sum(cat['amount'] for cat in category_data)
    for cat in category_data:
        percentage = (cat['amount'] / total * 100) if total > 0 else 0
        ws.append([
            cat['name'],
            cat['count'],
            cat['amount'],
            f"{percentage:.1f}%"
        ])

    # Total
    ws.append([])
    ws.append(['TOTAL', sum(cat['count'] for cat in category_data), total, '100%'])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_category_distribution_pdf(category_data, user):
    """Exporte la répartition par catégorie en PDF avec graphique"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Répartition des Abonnements & Crédits", user)

    if not category_data:
        elements.append(Paragraph("Aucune donnée disponible", getSampleStyleSheet()['Normal']))
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # Calculer le total
    total = sum(cat['amount'] for cat in category_data)

    # Créer le graphique en camembert
    drawing = Drawing(400, 300)
    pie = Pie()
    pie.x = 100
    pie.y = 50
    pie.width = 200
    pie.height = 200

    # Préparer les données
    pie.data = [cat['amount'] for cat in category_data]
    pie.labels = [f"{cat['name'][:20]}\n{cat['amount']:.2f}€" for cat in category_data]

    # Couleurs personnalisées basées sur les données
    pie_colors = []
    for cat in category_data:
        # Utiliser les couleurs des catégories si disponibles
        if 'color' in cat and cat['color']:
            try:
                pie_colors.append(colors.HexColor(cat['color']))
            except:
                pie_colors.append(colors.blue)
        else:
            pie_colors.append(colors.blue)

    pie.slices.strokeWidth = 0.5
    for i, color in enumerate(pie_colors):
        pie.slices[i].fillColor = color

    drawing.add(pie)
    elements.append(drawing)
    elements.append(Spacer(1, 20))

    # Ajouter un tableau récapitulatif sous le graphique
    data = [['Catégorie', 'Montant mensuel', 'Pourcentage']]

    for cat in category_data:
        percentage = (cat['amount'] / total * 100) if total > 0 else 0
        data.append([
            cat['name'][:30],
            f"{cat['amount']:.2f} €",
            f"{percentage:.1f}%"
        ])

    # Ligne de total
    data.append(['TOTAL', f"{total:.2f} €", '100%'])

    table = Table(data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_monthly_evolution_excel(monthly_data, user):
    """Exporte l'évolution des revenus et dépenses mensuelles en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Evolution mensuelle"

    # En-tête du document
    ws['A1'] = f"Evolution des revenus et des dépenses mensuelles - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Mois', 'Abonnements', 'Crédits', 'Revenus']
    ws.append([])
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    for month_data in monthly_data:
        ws.append([
            month_data['month'],
            round(month_data['subscriptions'], 2),
            round(month_data['credits'], 2),
            round(month_data['revenues'], 2)
        ])

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_monthly_evolution_pdf(monthly_data, user):
    """Exporte l'évolution des revenus et dépenses mensuelles en PDF avec graphique"""
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.widgets.markers import makeMarker

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Évolution des revenus et des dépenses mensuelles", user)

    if not monthly_data:
        elements.append(Paragraph("Aucune donnée disponible", getSampleStyleSheet()['Normal']))
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # Créer le graphique en lignes
    drawing = Drawing(500, 300)
    lc = HorizontalLineChart()
    lc.x = 50
    lc.y = 50
    lc.height = 200
    lc.width = 400

    # Préparer les données (3 séries : abonnements, crédits, revenus)
    subscriptions_data = [month['subscriptions'] for month in monthly_data]
    credits_data = [month['credits'] for month in monthly_data]
    revenues_data = [month['revenues'] for month in monthly_data]

    lc.data = [subscriptions_data, credits_data, revenues_data]

    # Configurer les catégories (labels des mois)
    # Afficher seulement certains labels pour éviter le chevauchement
    month_labels = [month['month'] for month in monthly_data]
    # Afficher 1 label sur 2 pour la lisibilité
    lc.categoryAxis.categoryNames = month_labels
    lc.categoryAxis.labels.angle = 45
    lc.categoryAxis.labels.fontSize = 7
    lc.categoryAxis.labels.boxAnchor = 'e'

    # Configurer l'axe des valeurs
    lc.valueAxis.valueMin = 0
    max_value = max(max(subscriptions_data + credits_data + revenues_data, default=0), 100)
    lc.valueAxis.valueMax = max_value * 1.1
    lc.valueAxis.valueStep = max_value / 5

    # Couleurs des lignes (bleu pour abonnements, jaune pour crédits, vert pour revenus)
    lc.lines[0].strokeColor = colors.HexColor('#0d6efd')  # Bleu
    lc.lines[0].strokeWidth = 2
    lc.lines[0].symbol = makeMarker('FilledCircle', size=4)

    lc.lines[1].strokeColor = colors.HexColor('#ffc107')  # Jaune
    lc.lines[1].strokeWidth = 2
    lc.lines[1].symbol = makeMarker('FilledSquare', size=4)

    lc.lines[2].strokeColor = colors.HexColor('#6f42c1')  # Vert
    lc.lines[2].strokeWidth = 2
    lc.lines[2].symbol = makeMarker('FilledDiamond', size=4)

    drawing.add(lc)

    # Ajouter une légende manuelle
    legend = Legend()
    legend.x = 50
    legend.y = 270
    legend.dx = 8
    legend.dy = 8
    legend.fontSize = 9
    legend.columnMaximum = 3
    legend.alignment = 'right'
    legend.colorNamePairs = [
        (colors.HexColor('#0d6efd'), 'Abonnements'),
        (colors.HexColor('#ffc107'), 'Crédits'),
        (colors.HexColor('#6f42c1'), 'Revenus')
    ]
    drawing.add(legend)

    elements.append(drawing)
    elements.append(Spacer(1, 20))

    # Ajouter un tableau récapitulatif sous le graphique
    table_data = [['Mois', 'Abonnements', 'Crédits', 'Revenus']]

    for month_data in monthly_data:
        table_data.append([
            month_data['month'][:15],
            f"{month_data['subscriptions']:.2f} €",
            f"{month_data['credits']:.2f} €",
            f"{month_data['revenues']:.2f} €"
        ])

    table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_subscriptions_excel(subscriptions, user):
    """Exporte la liste des abonnements en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Mes abonnements"

    # En-tête du document
    ws['A1'] = f"Mes abonnements - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Nom', 'Catégorie', 'Service', 'Montant', 'Devise', 'Cycle', 'Date début', 'Prochain paiement', 'Statut']
    ws.append([])
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    for sub in subscriptions:
        ws.append([
            sub.name,
            sub.category.name if sub.category else '-',
            sub.service.name if sub.service else '-',
            sub.amount,
            sub.currency,
            sub.billing_cycle,
            sub.start_date.strftime('%d/%m/%Y'),
            sub.next_billing_date.strftime('%d/%m/%Y'),
            'Actif' if sub.is_active else 'Inactif'
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_subscriptions_pdf(subscriptions, user):
    """Exporte la liste des abonnements en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Mes abonnements", user)

    # Table
    data = [['Nom', 'Catégorie', 'Montant', 'Cycle', 'Début', 'Prochain', 'Statut']]

    for sub in subscriptions:
        data.append([
            sub.name[:20],
            (sub.category.name[:15] if sub.category else '-'),
            f"{sub.amount:.2f}",
            sub.billing_cycle[:7],
            sub.start_date.strftime('%d/%m/%y'),
            sub.next_billing_date.strftime('%d/%m/%y'),
            'Actif' if sub.is_active else 'Inactif'
        ])

    table = Table(data, colWidths=[1.8*inch, 1.3*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_categories_excel(categories, user):
    """Exporte la liste des catégories en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Mes catégories"

    # En-tête du document
    ws['A1'] = f"Mes catégories - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Nom', 'Description', 'Couleur', 'Icône', 'Type', 'Date création']
    ws.append([])
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    for cat in categories:
        ws.append([
            cat.name,
            cat.description[:50] if cat.description else '-',
            cat.color,
            cat.icon if cat.icon else '-',
            'Personnalisée' if cat.user_id else 'Globale',
            cat.created_at.strftime('%d/%m/%Y')
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_categories_pdf(categories, user):
    """Exporte la liste des catégories en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Mes catégories", user)

    # Table
    data = [['Nom', 'Description', 'Type', 'Date création']]

    for cat in categories:
        data.append([
            cat.name[:25],
            (cat.description[:35] + '...' if cat.description and len(cat.description) > 35 else cat.description or '-'),
            'Perso.' if cat.user_id else 'Globale',
            cat.created_at.strftime('%d/%m/%Y')
        ])

    table = Table(data, colWidths=[1.8*inch, 2.5*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_services_excel(services, user):
    """Exporte la liste des services en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Mes services"

    # En-tête du document
    ws['A1'] = f"Mes services - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Nom', 'Catégorie', 'Description', 'Nb formules', 'Type', 'Date création']
    ws.append([])
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    for service in services:
        ws.append([
            service.name,
            service.category.name if service.category else '-',
            service.description[:50] if service.description else '-',
            len(service.plans),
            'Personnalisé' if service.user_id else 'Global',
            service.created_at.strftime('%d/%m/%Y')
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_services_pdf(services, user):
    """Exporte la liste des services en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Mes services", user)

    # Table
    data = [['Nom', 'Catégorie', 'Description', 'Formules', 'Type']]

    for service in services:
        data.append([
            service.name[:20],
            (service.category.name[:15] if service.category else '-'),
            (service.description[:35] + '...' if service.description and len(service.description) > 35 else service.description or '-'),
            str(len(service.plans)),
            'Perso.' if service.user_id else 'Global'
        ])

    table = Table(data, colWidths=[1.5*inch, 1.3*inch, 2.2*inch, 0.8*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_upcoming_credits_excel(credits, user):
    """Exporte les prochains prélèvements pour les crédits en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Crédits"

    # En-tête du document
    ws['A1'] = f"Crédits : Prochains prélèvements - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Date de paiement', 'Nom du crédit', 'Montant', 'Devise', 'Cycle', 'Montant restant', 'Taux intérêt', 'Jours restants']
    ws.append([])  # Ligne vide
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    now = datetime.now().date()
    for credit in credits:
        days_until = (credit.next_payment_date - now).days
        ws.append([
            credit.next_payment_date.strftime('%d/%m/%Y'),
            credit.name,
            credit.amount,
            credit.currency,
            credit.billing_cycle,
            credit.remaining_amount if credit.remaining_amount else '-',
            f"{credit.interest_rate}%" if credit.interest_rate else '-',
            days_until
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_upcoming_credits_pdf(credits, user):
    """Exporte les prochains prélèvements pour les crédits en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Crédits : Prochains prélèvements", user)

    # Table
    data = [['Date paiement', 'Nom du crédit', 'Montant', 'Cycle', 'Restant', 'Taux', 'Jours']]
    now = datetime.now().date()

    for credit in credits:
        days_until = (credit.next_payment_date - now).days
        remaining = f"{credit.remaining_amount:.2f}" if credit.remaining_amount else '-'
        interest = f"{credit.interest_rate}%" if credit.interest_rate else '-'

        data.append([
            credit.next_payment_date.strftime('%d/%m/%Y'),
            credit.name[:25],  # Limiter la longueur
            f"{credit.amount:.2f} {credit.currency}",
            credit.billing_cycle,
            remaining,
            interest,
            str(days_until)
        ])

    table = Table(data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 1*inch, 1.2*inch, 0.8*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_upcoming_revenues_excel(revenues, user):
    """Exporte les prochains versements pour les revenus en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Revenus"

    # En-tête du document
    ws['A1'] = f"Revenus : Prochains versements - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Date de paiement', 'Nom du revenu', 'Montant', 'Devise', 'Cycle', 'Employeur', 'Jours restants']
    ws.append([])  # Ligne vide
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    now = datetime.now().date()
    for revenue in revenues:
        days_until = (revenue.next_payment_date - now).days
        employer_name = revenue.employer.name if revenue.employer else '-'
        ws.append([
            revenue.next_payment_date.strftime('%d/%m/%Y'),
            revenue.name,
            revenue.amount,
            revenue.currency,
            revenue.billing_cycle,
            employer_name,
            days_until
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_upcoming_revenues_pdf(revenues, user):
    """Exporte les prochains versements pour les revenus en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Revenus : Prochains versements", user)

    # Table
    data = [['Date paiement', 'Nom du revenu', 'Montant', 'Cycle', 'Employeur', 'Jours']]
    now = datetime.now().date()

    for revenue in revenues:
        days_until = (revenue.next_payment_date - now).days
        employer_name = revenue.employer.name[:20] if revenue.employer else '-'

        data.append([
            revenue.next_payment_date.strftime('%d/%m/%Y'),
            revenue.name[:30],  # Limiter la longueur
            f"{revenue.amount:.2f} {revenue.currency}",
            revenue.billing_cycle,
            employer_name,
            str(days_until)
        ])

    table = Table(data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 1*inch, 1.5*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_revenue_distribution_excel(revenue_list, user):
    """Exporte la répartition des versements en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Répartition revenus"

    # En-tête du document
    ws['A1'] = f"Répartition des versements - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Employeur / Source', 'Montant mensuel (€)']
    ws.append([])  # Ligne vide
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    total = 0
    for item in revenue_list:
        ws.append([
            item['name'],
            round(item['total'], 2)
        ])
        total += item['total']

    # Ligne de total
    ws.append([])
    total_row = ws.max_row
    ws[f'A{total_row}'] = 'TOTAL'
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'B{total_row}'] = round(total, 2)
    ws[f'B{total_row}'].font = Font(bold=True, size=12)

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_revenue_distribution_pdf(revenue_list, user):
    """Exporte la répartition des versements en PDF avec graphique"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Répartition des versements", user)

    if not revenue_list:
        elements.append(Paragraph("Aucune donnée disponible", getSampleStyleSheet()['Normal']))
        doc.build(elements)
        buffer.seek(0)
        return buffer

    # Calculer le total
    total = sum(item['total'] for item in revenue_list)

    # Créer le graphique en camembert
    drawing = Drawing(400, 300)
    pie = Pie()
    pie.x = 100
    pie.y = 50
    pie.width = 200
    pie.height = 200

    # Préparer les données
    pie.data = [item['total'] for item in revenue_list]
    pie.labels = [f"{item['name'][:20]}\n{item['total']:.2f}€" for item in revenue_list]

    # Couleurs personnalisées (palette violette pour les revenus)
    colors_palette = [
        colors.HexColor('#6f42c1'),
        colors.HexColor('#8b5cf6'),
        colors.HexColor('#a78bfa'),
        colors.HexColor('#c4b5fd'),
        colors.HexColor('#ddd6fe'),
        colors.HexColor('#ede9fe')
    ]

    pie.slices.strokeWidth = 0.5
    for i in range(len(revenue_list)):
        pie.slices[i].fillColor = colors_palette[i % len(colors_palette)]

    drawing.add(pie)
    elements.append(drawing)
    elements.append(Spacer(1, 20))

    # Ajouter un tableau récapitulatif sous le graphique
    data = [['Employeur / Source', 'Montant mensuel', 'Pourcentage']]

    for item in revenue_list:
        percentage = (item['total'] / total * 100) if total > 0 else 0
        data.append([
            item['name'][:30],
            f"{item['total']:.2f} €",
            f"{percentage:.1f}%"
        ])

    # Ligne de total
    data.append(['TOTAL', f"{total:.2f} €", '100%'])

    table = Table(data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E2F3')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def export_unpointed_checks_excel(checks, user):
    """Exporte les chèques non débités en Excel"""
    wb = create_excel_workbook()
    ws = wb.active
    ws.title = "Chèques non débités"

    # En-tête du document
    ws['A1'] = f"Chèques : Non débités - {user.first_name} {user.last_name or ''}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)

    # En-têtes des colonnes
    headers = ['Date', 'Numéro de chèque', 'Bénéficiaire', 'Montant', 'Devise', 'Description']
    ws.append([])  # Ligne vide
    ws.append(headers)
    style_excel_header(ws, row=4)

    # Données
    for check in checks:
        ws.append([
            check.transaction_date.strftime('%d/%m/%Y'),
            check.name,
            check.description or '-',
            check.amount,
            check.currency,
            check.description or ''
        ])

    # Ajuster la largeur des colonnes
    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 20

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_unpointed_checks_pdf(checks, user):
    """Exporte les chèques non débités en PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter l'en-tête avec logo
    add_pdf_header(elements, "Chèques : Non débités", user)

    # Table
    data = [['Date', 'Numéro de chèque', 'Bénéficiaire', 'Montant']]

    for check in checks:
        data.append([
            check.transaction_date.strftime('%d/%m/%Y'),
            check.name[:40],  # Limiter la longueur
            (check.description or '-')[:30],
            f"{check.amount:.2f} {check.currency}"
        ])

    table = Table(data, colWidths=[1.2*inch, 2.5*inch, 2*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC3545')),  # Rouge pour les dépenses
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
